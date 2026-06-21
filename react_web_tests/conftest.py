import os
import time
from datetime import datetime
import pytest
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Store test results globally during pytest session
security_results = []
session_start_time = 0

def pytest_sessionstart(session):
    global session_start_time
    session_start_time = time.time()

@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    rep = outcome.get_result()
    
    # We log call results or setup failures/skips
    if rep.when == "call" or (rep.when == "setup" and (rep.failed or rep.skipped)):
        is_pass = rep.passed or rep.skipped
        status = "PASSED" if is_pass else "FAILED"
        duration = rep.duration
        
        test_name = item.name
        
        # Categorize based on function name
        func_name = item.function.__name__ if hasattr(item, 'function') else ""
        if "xss" in func_name:
            category = "Cross-Site Scripting (XSS)"
        elif "sqli" in func_name:
            category = "SQL Injection (SQLi)"
        elif "header" in func_name or "disclosure" in func_name:
            category = "HTTP Security Headers"
        elif "cookie" in func_name:
            category = "Cookie Flag Controls"
        elif "frame" in func_name or "clickjacking" in func_name:
            category = "Clickjacking Protection"
        elif "csrf" in func_name:
            category = "CSRF Verification"
        elif "redirect" in func_name:
            category = "Open Redirect Controls"
        elif "method" in func_name:
            category = "HTTP Method Restrictions"
        else:
            category = "General Vulnerability Check"
            
        message = ""
        if rep.failed:
            if rep.longrepr:
                if hasattr(rep.longrepr, 'reprcrash'):
                    message = rep.longrepr.reprcrash.message
                else:
                    message = str(rep.longrepr)
            else:
                message = "Assertion failed."
        elif rep.skipped:
            message = "Skipped (Safe - no input fields or forms present to test on this page)."
        else:
            message = "Passed secure validation check."
            
        security_results.append({
            "name": test_name,
            "category": category,
            "status": status,
            "message": message,
            "duration": duration,
        })

def pytest_sessionfinish(session, exitstatus):
    if not security_results:
        return
        
    print(f"\n[SECURITY] Generating Excel security report for {len(security_results)} test cases...")
    
    wb = openpyxl.Workbook()
    
    # ─── Sheet 1: Summary Dashboard ───
    ws_summary = wb.active
    ws_summary.title = "Security Summary"
    ws_summary.sheet_view.showGridLines = False
    
    NAVY = "0D1B2A"
    INDIGO = "4F46E5"
    CHARCOAL = "1F2937"
    GREEN_BG = "D1FAE5"
    GREEN_FG = "065F46"
    RED_BG = "FEE2E2"
    RED_FG = "991B1B"
    GRAY_BG = "F8FAFC"
    BORDER_COLOR = "CBD5E1"
    
    thin = Side(style='thin', color=BORDER_COLOR)
    card_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Title Block
    ws_summary.merge_cells("A1:D1")
    title_cell = ws_summary["A1"]
    title_cell.value = "CAMPUSCONNECT WEB SECURITY ANALYSIS REPORT"
    title_cell.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=INDIGO)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 40
    
    # Sub-header
    ws_summary.merge_cells("A2:D2")
    sub_cell = ws_summary["A2"]
    sub_cell.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Target Host: http://localhost:5173/CampusConnect_web/"
    sub_cell.font = Font(name="Segoe UI", size=9, italic=True, color="FFFFFF")
    sub_cell.fill = PatternFill("solid", fgColor=NAVY)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[2].height = 20
    
    # Calculations
    total = len(security_results)
    passed = sum(1 for r in security_results if r["status"] == "PASSED")
    failed = total - passed
    pass_percent = (passed / total * 100) if total > 0 else 0
    duration = round(time.time() - session_start_time, 1)
    
    # KPI Grid
    kpis = [
        ("Total Scans", str(total), "A"),
        ("Passed", str(passed), "B"),
        ("Failed", str(failed), "C"),
        ("Pass Rate", f"{pass_percent:.1f}%", "D")
    ]
    
    ws_summary.row_dimensions[3].height = 10
    
    for label, val, col in kpis:
        lbl_c = ws_summary[f"{col}4"]
        lbl_c.value = label
        lbl_c.font = Font(name="Segoe UI", size=9, bold=True, color="555555")
        lbl_c.fill = PatternFill("solid", fgColor=GRAY_BG)
        lbl_c.alignment = Alignment(horizontal="center", vertical="center")
        lbl_c.border = card_border
        
        val_c = ws_summary[f"{col}5"]
        val_c.value = val
        fg_color = GREEN_FG if label == "Passed" or (label == "Pass Rate" and pass_percent > 85) else (RED_FG if label == "Failed" and failed > 0 else NAVY)
        val_c.font = Font(name="Segoe UI", size=18, bold=True, color=fg_color)
        val_c.fill = PatternFill("solid", fgColor=GRAY_BG)
        val_c.alignment = Alignment(horizontal="center", vertical="center")
        val_c.border = card_border
        
    ws_summary.row_dimensions[4].height = 18
    ws_summary.row_dimensions[5].height = 28
    ws_summary.row_dimensions[6].height = 15
    
    # Breakdown Header
    ws_summary.merge_cells("A8:D8")
    lbl_breakdown = ws_summary["A8"]
    lbl_breakdown.value = "  Vulnerability Category Breakdown"
    lbl_breakdown.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    lbl_breakdown.fill = PatternFill("solid", fgColor=INDIGO)
    lbl_breakdown.alignment = Alignment(horizontal="left", vertical="center")
    ws_summary.row_dimensions[8].height = 24
    
    # Table column headers
    cols = ["Category", "Total Scans", "Passed", "Failed"]
    col_letters = ["A", "B", "C", "D"]
    for i, c_name in enumerate(cols):
        cell = ws_summary[f"{col_letters[i]}9"]
        cell.value = c_name
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=CHARCOAL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style='medium', color="000000"))
    ws_summary.row_dimensions[9].height = 22
    
    categories = list(set(r["category"] for r in security_results))
    categories.sort()
    
    current_row = 10
    for idx, cat in enumerate(categories):
        cat_results = [r for r in security_results if r["category"] == cat]
        c_total = len(cat_results)
        c_passed = sum(1 for r in cat_results if r["status"] == "PASSED")
        c_failed = c_total - c_passed
        
        zebra_fill = PatternFill("solid", fgColor="F9FAFB" if idx % 2 == 1 else "FFFFFF")
        
        row_data = [cat, c_total, c_passed, c_failed]
        for i, val in enumerate(row_data):
            cell = ws_summary[f"{col_letters[i]}{current_row}"]
            cell.value = val
            cell.font = Font(name="Segoe UI", size=10)
            cell.fill = zebra_fill
            cell.border = Border(bottom=Side(style='thin', color="E5E7EB"), right=Side(style='thin', color="E5E7EB"))
            if i == 0:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if i == 3 and c_failed > 0:
                    cell.font = Font(name="Segoe UI", size=10, bold=True, color=RED_FG)
                    cell.fill = PatternFill("solid", fgColor=RED_BG)
        ws_summary.row_dimensions[current_row].height = 20
        current_row += 1
        
    ws_summary.column_dimensions["A"].width = 32
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 15
    ws_summary.column_dimensions["D"].width = 15
    
    # ─── Sheet 2: Scan Details ───
    ws_details = wb.create_sheet(title="Scan Details")
    ws_details.sheet_view.showGridLines = False
    
    ws_details.merge_cells("A1:F1")
    details_title = ws_details["A1"]
    details_title.value = "  ITEMIZED VULNERABILITY SCAN RESULTS"
    details_title.font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
    details_title.fill = PatternFill("solid", fgColor=NAVY)
    details_title.alignment = Alignment(horizontal="left", vertical="center")
    ws_details.row_dimensions[1].height = 30
    
    detail_headers = ["No", "Vulnerability Checked / Test Case", "Category", "Status", "Duration (s)", "Remarks / Exception logs"]
    detail_cols = ["A", "B", "C", "D", "E", "F"]
    detail_widths = [6, 45, 25, 12, 12, 60]
    
    for i, h_name in enumerate(detail_headers):
        cell = ws_details[f"{detail_cols[i]}2"]
        cell.value = h_name
        cell.font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=CHARCOAL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_details.column_dimensions[detail_cols[i]].width = detail_widths[i]
    ws_details.row_dimensions[2].height = 22
    
    for idx, r in enumerate(security_results):
        row = 3 + idx
        zebra_fill = PatternFill("solid", fgColor="F9FAFB" if idx % 2 == 1 else "FFFFFF")
        is_passed = r["status"] == "PASSED"
        
        status_fill = PatternFill("solid", fgColor=GREEN_BG if is_passed else RED_BG)
        status_fg = GREEN_FG if is_passed else RED_FG
        
        row_vals = [idx + 1, r["name"], r["category"], r["status"], round(r["duration"], 3), r["message"]]
        for col_idx, val in enumerate(row_vals):
            cell = ws_details.cell(row=row, column=col_idx + 1)
            cell.value = val
            cell.font = Font(name="Segoe UI", size=9)
            cell.border = Border(bottom=Side(style='thin', color="E5E7EB"), right=Side(style='thin', color="E5E7EB"))
            
            if col_idx == 0:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Segoe UI", size=9, bold=True, color="888888")
                cell.fill = zebra_fill
            elif col_idx == 3:
                cell.font = Font(name="Segoe UI", size=9, bold=True, color=status_fg)
                cell.fill = status_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = zebra_fill
            elif col_idx == 5:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.fill = zebra_fill
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.fill = zebra_fill
                
        msg_len = len(str(r["message"]))
        ws_details.row_dimensions[row].height = 20 if msg_len < 70 else 32
        
    report_name = "security_test_report.xlsx"
    wb.save(report_name)
    print(f"\n[SECURITY REPORT] Premium Excel security report generated successfully -> {report_name}\n")

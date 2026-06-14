"""
excel_reporter.py — Premium Excel report generator for React E2E test results.

Generates a multi-sheet workbook:
  Sheet 1: Summary Dashboard (KPI cards + category breakdown)
  Sheets 2-9: Per-category detail tables (colour-coded PASSED/FAILED rows)
"""
import os
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint


# ── Palette ──────────────────────────────────────────────────────────────────
CLR = {
    "deep_navy":   "0D1B2A",
    "brand_blue":  "1A56DB",
    "brand_light": "3B82F6",
    "accent_teal": "0EA5E9",
    "hdr_bg":      "1E3A5F",
    "passed_bg":   "D1FAE5",
    "passed_fg":   "065F46",
    "failed_bg":   "FEE2E2",
    "failed_fg":   "991B1B",
    "zebra":       "F0F7FF",
    "white":       "FFFFFF",
    "light_gray":  "F8FAFC",
    "border":      "CBD5E1",
    "subtitle_bg": "EFF6FF",
    "chart_bar1":  "1A56DB",
    "chart_bar2":  "10B981",
}

FF  = "Calibri"  # font family


def _border(color=CLR["border"]):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(size=10, bold=False, color=CLR["deep_navy"], italic=False):
    return Font(name=FF, size=size, bold=bold, color=color, italic=italic)


def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Category Registry ────────────────────────────────────────────────────────
CATEGORIES = [
    {"name": "Authentication",        "prefix": "FN-AUTH", "sheet": "Authentication",       "icon": "🔐"},
    {"name": "Explore Freelancers",   "prefix": "FN-EXP",  "sheet": "Explore Freelancers",  "icon": "🔍"},
    {"name": "Gigs Board",            "prefix": "FN-GIG",  "sheet": "Gigs Board",           "icon": "💼"},
    {"name": "Teams",                 "prefix": "FN-TEAM", "sheet": "Teams",                "icon": "👥"},
    {"name": "Messages / Chats",      "prefix": "FN-MSG",  "sheet": "Messages & Chats",     "icon": "💬"},
    {"name": "Dashboard",             "prefix": "FN-DASH", "sheet": "Dashboard",            "icon": "📊"},
    {"name": "Profile & Bookmarks",   "prefix": "FN-PROF", "sheet": "Profile & Bookmarks",  "icon": "👤"},
    {"name": "Navigation & Routing",  "prefix": "FN-NAV",  "sheet": "Navigation & Routing", "icon": "🧭"},
]


class ExcelReporter:
    def __init__(self):
        self.results = []
        self.start_time = time.time()

    def add_result(self, test_name, status, message="", duration=0.0):
        self.results.append({
            "name":      test_name,
            "status":    status,
            "duration":  round(duration, 3),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "message":   message,
        })

    # ─── Public Entry Point ───────────────────────────────────────────────────

    def save(self, filepath):
        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)

        wb = Workbook()
        self._build_summary_sheet(wb.active)
        for cat in CATEGORIES:
            self._build_detail_sheet(wb.create_sheet(title=cat["sheet"][:31]), cat)

        wb.save(filepath)
        print(f"\n[SAVED] Excel report saved -> {filepath}")
        return filepath

    # ─── Summary Dashboard Sheet ──────────────────────────────────────────────

    def _build_summary_sheet(self, ws):
        ws.title = "Summary Dashboard"
        ws.sheet_view.showGridLines = False

        total  = len(self.results)
        passed = sum(1 for r in self.results if r["status"] == "PASSED")
        failed = total - passed
        rate   = (passed / total * 100) if total > 0 else 0
        dur    = round(time.time() - self.start_time, 1)

        # ── Title banner ──────────────────────────────────────────────────────
        ws.merge_cells("A1:J1")
        c = ws["A1"]
        c.value       = "  CampusConnect React Web Platform — Selenium E2E Test Report"
        c.font        = _font(18, bold=True, color=CLR["white"])
        c.fill        = _fill(CLR["hdr_bg"])
        c.alignment   = _align("left", "center")
        ws.row_dimensions[1].height = 44

        # sub-header
        ws.merge_cells("A2:J2")
        c = ws["A2"]
        c.value     = f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   App URL: http://localhost:5173/CampusConnect_web/"
        c.font      = _font(9, italic=True, color=CLR["accent_teal"])
        c.fill      = _fill(CLR["deep_navy"])
        c.alignment = _align("left", "center")
        ws.row_dimensions[2].height = 18

        # ── KPI Cards Row ─────────────────────────────────────────────────────
        ws.row_dimensions[3].height = 10  # spacer

        kpis = [
            ("Total Tests", str(total),     CLR["brand_blue"],  "A"),
            ("Passed",      str(passed),    CLR["passed_fg"],   "C"),
            ("Failed",      str(failed),    CLR["failed_fg"],   "E"),
            ("Pass Rate",   f"{rate:.1f}%", CLR["brand_blue"],  "G"),
            ("Duration",    f"{dur}s",      CLR["deep_navy"],   "I"),
        ]

        for label, val, fg, col in kpis:
            self._kpi_card(ws, col, label, val, fg)

        ws.row_dimensions[4].height = 14
        ws.row_dimensions[5].height = 28
        ws.row_dimensions[6].height = 12

        # ── Category Breakdown Table ──────────────────────────────────────────
        header_row = 8
        ws.row_dimensions[7].height = 10  # spacer

        ws.merge_cells(f"A{header_row}:J{header_row}")
        c = ws[f"A{header_row}"]
        c.value     = "  Category Breakdown"
        c.font      = _font(12, bold=True, color=CLR["white"])
        c.fill      = _fill(CLR["brand_blue"])
        c.alignment = _align("left", "center")
        ws.row_dimensions[header_row].height = 28

        col_headers = ["", "Test Category", "Total Cases", "Passed", "Failed", "Pass %", "Status"]
        col_widths  = [3, 30, 14, 10, 10, 10, 16]
        col_letters = ["A", "B", "C", "D", "E", "F", "G"]

        sub_row = header_row + 1
        for i, (hdr, w) in enumerate(zip(col_headers, col_widths)):
            cell = ws[f"{col_letters[i]}{sub_row}"]
            cell.value     = hdr
            cell.font      = _font(10, bold=True, color=CLR["deep_navy"])
            cell.fill      = _fill(CLR["subtitle_bg"])
            cell.border    = _border()
            cell.alignment = _align("center")
            ws.column_dimensions[col_letters[i]].width = w
        ws.row_dimensions[sub_row].height = 22

        data_row = sub_row + 1
        for i, cat in enumerate(CATEGORIES):
            prefix    = cat["prefix"]
            cat_res   = [r for r in self.results if r["name"].startswith(prefix)]
            c_total   = len(cat_res)
            c_passed  = sum(1 for r in cat_res if r["status"] == "PASSED")
            c_failed  = c_total - c_passed
            c_rate    = (c_passed / c_total * 100) if c_total > 0 else 0
            row_fill  = _fill(CLR["zebra"]) if i % 2 == 0 else _fill(CLR["white"])
            status    = "✅ PASS" if c_failed == 0 else "❌ FAIL"

            vals = [cat["icon"], cat["name"], c_total, c_passed, c_failed, f"{c_rate:.0f}%", status]
            for j, (letter, val) in enumerate(zip(col_letters, vals)):
                cell = ws[f"{letter}{data_row}"]
                cell.value     = val
                cell.font      = _font(10, color=CLR["passed_fg"] if "✅" in str(val) else
                                       CLR["failed_fg"] if "❌" in str(val) else CLR["deep_navy"])
                cell.fill      = row_fill
                cell.border    = _border()
                cell.alignment = _align("center" if j != 1 else "left")
            ws.row_dimensions[data_row].height = 20
            data_row += 1

        # ── Totals row ────────────────────────────────────────────────────────
        for letter, val in zip(col_letters, ["", "TOTAL", total, passed, failed, f"{rate:.0f}%", ""]):
            cell = ws[f"{letter}{data_row}"]
            cell.value     = val
            cell.font      = _font(10, bold=True, color=CLR["deep_navy"])
            cell.fill      = _fill(CLR["subtitle_bg"])
            cell.border    = _border()
            cell.alignment = _align("center")
        ws.row_dimensions[data_row].height = 22

        # widen A col
        ws.column_dimensions["A"].width = 4
        # widen remaining
        for let, w in [("H", 3), ("I", 14), ("J", 14)]:
            ws.column_dimensions[let].width = w

    def _kpi_card(self, ws, col_letter, label, value, fg_color):
        """Render a 2-row merged KPI card in the summary sheet."""
        label_row = 4
        value_row = 5

        # label cell
        cell_l = ws[f"{col_letter}{label_row}"]
        cell_l.value     = label
        cell_l.font      = _font(9, bold=True, color="888888")
        cell_l.fill      = _fill(CLR["light_gray"])
        cell_l.alignment = _align("center", "center")
        cell_l.border    = _border(CLR["border"])

        # value cell
        cell_v = ws[f"{col_letter}{value_row}"]
        cell_v.value     = value
        cell_v.font      = _font(20, bold=True, color=fg_color)
        cell_v.fill      = _fill(CLR["light_gray"])
        cell_v.alignment = _align("center", "center")
        cell_v.border    = _border(CLR["border"])

        # Width
        ws.column_dimensions[col_letter].width = 14

    # ─── Detail Sheet per Category ────────────────────────────────────────────

    def _build_detail_sheet(self, ws, cat):
        ws.sheet_view.showGridLines = False
        prefix  = cat["prefix"]
        icon    = cat["icon"]
        name    = cat["name"]
        results = [r for r in self.results if r["name"].startswith(prefix)]

        total  = len(results)
        passed = sum(1 for r in results if r["status"] == "PASSED")
        failed = total - passed

        # ── Title banner ──────────────────────────────────────────────────────
        ws.merge_cells("A1:G1")
        c = ws["A1"]
        c.value     = f"  {icon}  CampusConnect — {name} Tests"
        c.font      = _font(14, bold=True, color=CLR["white"])
        c.fill      = _fill(CLR["hdr_bg"])
        c.alignment = _align("left", "center")
        ws.row_dimensions[1].height = 36

        # Summary strip
        ws.merge_cells("A2:G2")
        c = ws["A2"]
        c.value     = (f"  Total: {total}   |   Passed: {passed}   |   Failed: {failed}   |   "
                       f"Pass Rate: {(passed/total*100):.0f}%" if total else "  No results.")
        c.font      = _font(9, color=CLR["white"])
        c.fill      = _fill(CLR["brand_blue"])
        c.alignment = _align("left", "center")
        ws.row_dimensions[2].height = 16

        # ── Column headers ────────────────────────────────────────────────────
        headers = ["#", "Test ID & Name", "Status", "Duration (s)", "Timestamp", "Remarks / Details"]
        widths  = [5, 40, 12, 14, 22, 55]
        for col, (hdr, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=3, column=col)
            cell.value     = hdr
            cell.font      = _font(10, bold=True, color=CLR["white"])
            cell.fill      = _fill(CLR["hdr_bg"])
            cell.border    = _border()
            cell.alignment = _align("center")
            ws.column_dimensions[get_column_letter(col)].width = w

        # Merge remarks header across F-G
        ws.merge_cells("F3:G3")
        ws.row_dimensions[3].height = 22

        # ── Data rows ─────────────────────────────────────────────────────────
        for idx, r in enumerate(results):
            row = 4 + idx
            is_zebra   = idx % 2 == 1
            row_bg     = _fill(CLR["zebra"]) if is_zebra else _fill(CLR["white"])
            is_passed  = r["status"] == "PASSED"
            status_bg  = _fill(CLR["passed_bg"]) if is_passed else _fill(CLR["failed_bg"])
            status_fg  = CLR["passed_fg"]         if is_passed else CLR["failed_fg"]

            values = [idx + 1, r["name"], r["status"], r["duration"], r["timestamp"], r["message"]]
            for col, val in enumerate(values, 1):
                if col == 6 and col == 6:
                    # Merge F-G for remarks
                    ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
                cell = ws.cell(row=row, column=col)
                cell.value  = val
                cell.border = _border()

                if col == 3:  # Status column
                    cell.font      = _font(10, bold=True, color=status_fg)
                    cell.fill      = status_bg
                    cell.alignment = _align("center")
                elif col == 1:
                    cell.font      = _font(9, bold=True, color="888888")
                    cell.fill      = row_bg
                    cell.alignment = _align("center")
                elif col == 4:
                    cell.font      = _font(9, color=CLR["deep_navy"])
                    cell.fill      = row_bg
                    cell.alignment = _align("center")
                else:
                    cell.font      = _font(9, color=CLR["deep_navy"])
                    cell.fill      = row_bg
                    cell.alignment = _align("left", "center", wrap=(col == 6))

            ws.row_dimensions[row].height = 22 if len(str(r["message"])) < 80 else 36
